from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.urls import reverse, reverse_lazy
from django.views.generic import View,TemplateView,ListView,UpdateView,CreateView,DeleteView,DetailView, RedirectView
from django.contrib.auth.views import LoginView, LogoutView
from django.core.paginator import Paginator


from .models import Configuracion, Producto, Puerto, Pais
from django.contrib.auth.mixins import LoginRequiredMixin
from .form import *


from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders

from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.utils import *
from openpyxl.workbook.protection import WorkbookProtection

class Home(TemplateView):
    template_name = 'base.html'

class UserLoginView(LoginView):
    template_name = 'login.html' 
    redirect_authenticated_user = False
    success_url = reverse_lazy('base')  
    
class UserLogoutView(LogoutView):
    next_page = reverse_lazy('login')

""" CRUD producto """
class ProductoListView(LoginRequiredMixin, ListView):
    model = Producto
    form_class = ProductoForm
    template_name = 'producto/lista_producto.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(referencia__icontains=query)| queryset.filter(descripcion__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Productos', 'url': '/productos/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response

class ProductoUpdateView(LoginRequiredMixin,UpdateView):
    model = Producto
    template_name = 'producto/actualizar_producto.html'
    form_class = ProductoForm
    success_url = reverse_lazy('productos_listado')

class ProductoDeleteView(LoginRequiredMixin,DeleteView):
    model = Producto
    template_name = 'producto/eliminar_producto.html'
    success_url = reverse_lazy('productos_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        producto = self.get_object()
        producto.estado = False
        producto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('productos_listado')

class ProductoCreateView(LoginRequiredMixin,CreateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'producto/crear_producto.html'
    success_url = reverse_lazy('productos_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('productos_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
""" CRUD puerto """    
class PuertoListView(LoginRequiredMixin, ListView):
    model = Puerto
    form_class = PuertoForm
    template_name = 'puerto/lista_puerto.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre_puerto__icontains=query)| queryset.filter(sigla__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Puerto', 'url': '/puertos/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class PuertoCreateView(LoginRequiredMixin,CreateView):
    model = Puerto
    form_class = PuertoForm
    template_name = 'puerto/crear_puerto.html'
    success_url = reverse_lazy('puerto_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('puerto_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class PuertoUpdateView(LoginRequiredMixin,UpdateView):
    model = Puerto
    template_name = 'puerto/actualizar_puerto.html'
    form_class = PuertoForm
    success_url = reverse_lazy('puerto_listado')
    
class PuertoDeleteView(LoginRequiredMixin,DeleteView):
    model = Puerto
    template_name = 'puerto/eliminar_puerto.html'
    success_url = reverse_lazy('puerto_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('puerto_listado')

""" CRUD pais """  
class PaisListView(LoginRequiredMixin, ListView):
    model = Pais
    form_class = PuertoForm
    template_name = 'pais/lista_pais.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre_pais__icontains=query)| queryset.filter(sigla__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'País', 'url': '/país/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class PaisCreateView(LoginRequiredMixin,CreateView):
    model = Pais
    form_class = PaisForm
    template_name = 'pais/crear_pais.html'
    success_url = reverse_lazy('pais_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('pais_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class PaisUpdateView(LoginRequiredMixin,UpdateView):
    model = Pais
    template_name = 'pais/actualizar_pais.html'
    form_class = PaisForm
    success_url = reverse_lazy('pais_listado')       

class PaisDeleteView(LoginRequiredMixin,DeleteView):
    model = Pais
    template_name = 'pais/eliminar_pais.html'
    success_url = reverse_lazy('pais_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('puerto_listado')      
    
""" CRUD proveedor """  
class ProveedorListView(LoginRequiredMixin, ListView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'proveedor/lista_proveedor.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre_proveedor__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Proveedor', 'url': '/proveedor/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class ProveedorCreateView(LoginRequiredMixin,CreateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'proveedor/crear_proveedor.html'
    success_url = reverse_lazy('proveedor_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('proveedor_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class ProveedorUpdateView(LoginRequiredMixin,UpdateView):
    model = Proveedor
    template_name = 'proveedor/actualizar_proveedor.html'
    form_class = ProveedorForm
    success_url = reverse_lazy('proveedor_listado')       

class ProveedorDeleteView(LoginRequiredMixin,DeleteView):
    model = Proveedor
    template_name = 'proveedor/eliminar_proveedor.html'
    success_url = reverse_lazy('proveedor_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('proveedor_listado')      
        
""" CRUD cliente """  
class ClienteListView(LoginRequiredMixin, ListView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'cliente/lista_cliente.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre_cliente__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Clientes', 'url': '/clientes/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class ClienteCreateView(LoginRequiredMixin,CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'cliente/crear_cliente.html'
    success_url = reverse_lazy('cliente_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('cliente_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class ClienteUpdateView(LoginRequiredMixin,UpdateView):
    model = Cliente
    template_name = 'cliente/actualizar_cliente.html'
    form_class = ClienteForm
    success_url = reverse_lazy('cliente_listado')       

class ClienteDeleteView(LoginRequiredMixin,DeleteView):
    model = Cliente
    template_name = 'cliente/eliminar_cliente.html'
    success_url = reverse_lazy('cliente_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('cliente_listado')
    
""" CRUD empresa """  
class EmpresaListView(LoginRequiredMixin, ListView):
    model = Empresa
    template_name = 'empresa/lista_empresa.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre_empresa__icontains=query) | queryset.filter(nombre_encargado__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Empresa', 'url': '/empresa/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = EmpresaForm(data=request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class EmpresaCreateView(LoginRequiredMixin, CreateView):
    model = Empresa
    form_class = EmpresaForm
    template_name = 'empresa/crear_empresa.html'
    success_url = reverse_lazy('empresa_listado')

    def form_valid(self, form):
        self.object = form.save()
        mensaje = f'{self.model.__name__} registrado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('empresa_listado')
    def form_invalid(self, form):
        mensaje = f'{self.model.__name__} no se ha podido registrar!'
        error = form.errors
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 400
        return response
        
class EmpresaUpdateView(LoginRequiredMixin,UpdateView):
    model = Empresa
    template_name = 'empresa/actualizar_empresa.html'
    form_class = EmpresaForm
    success_url = reverse_lazy('empresa_listado')       

class EmpresaDeleteView(LoginRequiredMixin,DeleteView):
    model = Empresa
    template_name = 'empresa/eliminar_empresa.html'
    success_url = reverse_lazy('empresa_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('empresa_listado')      
    
""" CRUD importadora """  
class ImportadoraListView(LoginRequiredMixin, ListView):
    model = Importadora
    form_class = ImportadoraForm
    template_name = 'importadora/lista_importadora.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre_importadora__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Importadora', 'url': '/importadora/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class ImportadoraCreateView(LoginRequiredMixin,CreateView):
    model = Importadora
    form_class = ImportadoraForm
    template_name = 'importadora/crear_importadora.html'
    success_url = reverse_lazy('importadora_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('importadora_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class ImportadoraUpdateView(LoginRequiredMixin,UpdateView):
    model = Importadora
    template_name = 'importadora/actualizar_importadora.html'
    form_class = ImportadoraForm
    success_url = reverse_lazy('importadora_listado')       

class ImportadoraDeleteView(LoginRequiredMixin,DeleteView):
    model = Importadora
    template_name = 'importadora/eliminar_importadora.html'
    success_url = reverse_lazy('importadora_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('importadora_listado')
    
""" CRUD unidad """  
class UnidadDeMedidaListView(LoginRequiredMixin, ListView):
    model = UnidadDeMedida
    form_class = UnidadDeMedidaForm
    template_name = 'unidad/lista_unidad.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Unidad de Medida', 'url': '/unidad/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class UnidadDeMedidaCreateView(LoginRequiredMixin,CreateView):
    model = UnidadDeMedida
    form_class = UnidadDeMedidaForm
    template_name = 'unidad/crear_unidad.html'
    success_url = reverse_lazy('unidad_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('unidad_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class UnidadDeMedidaUpdateView(LoginRequiredMixin,UpdateView):
    model = UnidadDeMedida
    template_name = 'unidad/actualizar_unidad.html'
    form_class = UnidadDeMedidaForm
    success_url = reverse_lazy('unidad_listado')       

class UnidadDeMedidaDeleteView(LoginRequiredMixin,DeleteView):
    model = UnidadDeMedida
    template_name = 'unidad/eliminar_unidad.html'
    success_url = reverse_lazy('unidad_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('unidad_listado')
    
""" CRUD partida """  
class PartidaArancelariaListView(LoginRequiredMixin, ListView):
    model = PartidaArancelaria
    form_class = PartidaArancelariaForm
    template_name = 'partida/lista_partida.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(partida__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Partida Arancelaria', 'url': '/partida/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class PartidaArancelariaCreateView(LoginRequiredMixin,CreateView):
    model = PartidaArancelaria
    form_class = PartidaArancelariaForm
    template_name = 'partida/crear_partida.html'
    success_url = reverse_lazy('partida_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('partida_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class PartidaArancelariaUpdateView(LoginRequiredMixin,UpdateView):
    model = PartidaArancelaria
    template_name = 'partida/actualizar_partida.html'
    form_class = PartidaArancelariaForm
    success_url = reverse_lazy('partida_listado')       

class PartidaArancelariaDeleteView(LoginRequiredMixin,DeleteView):
    model = PartidaArancelaria
    template_name = 'partida/eliminar_partida.html'
    success_url = reverse_lazy('partida_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('partida_listado')

""" CRUD estado """  
class EstadoListView(LoginRequiredMixin, ListView):
    model = Estado
    form_class = EstadoForm
    template_name = 'estado/lista_estado.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre_estado__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Estado', 'url': '/estado/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class EstadoCreateView(LoginRequiredMixin,CreateView):
    model = Estado
    form_class = EstadoForm
    template_name = 'estado/crear_estado.html'
    success_url = reverse_lazy('estado_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('estado_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class EstadoUpdateView(LoginRequiredMixin,UpdateView):
    model = Estado
    template_name = 'estado/actualizar_estado.html'
    form_class = EstadoForm
    success_url = reverse_lazy('estado_listado')       

class EstadoDeleteView(LoginRequiredMixin,DeleteView):
    model = Estado
    template_name = 'estado/eliminar_estado.html'
    success_url = reverse_lazy('estado_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('estado_listado')

""" CRUD transporte """  
class TransporteListView(LoginRequiredMixin, ListView):
    model = Transporte
    form_class = TransporteForm
    template_name = 'transporte/lista_transporte.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(nombre_transporte__icontains=query)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Tipo de Transporte', 'url': '/transporte/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class TransporteCreateView(LoginRequiredMixin,CreateView):
    model = Transporte
    form_class = TransporteForm
    template_name = 'transporte/crear_transporte.html'
    success_url = reverse_lazy('transporte_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('transporte_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class TransporteUpdateView(LoginRequiredMixin,UpdateView):
    model = Transporte
    template_name = 'transporte/actualizar_transporte.html'
    form_class = TransporteForm
    success_url = reverse_lazy('transporte_listado')       

class TransporteDeleteView(LoginRequiredMixin,DeleteView):
    model = Transporte
    template_name = 'transporte/eliminar_transporte.html'
    success_url = reverse_lazy('transporte_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        puerto = self.get_object()
        puerto.estado = False
        puerto.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('transporte_listado')

""" CRUD contenedor """  
class ContenedorListView(LoginRequiredMixin, ListView):
    model = Contenedor
    form_class = ContenedorForm
    template_name = 'contenedor/lista_contenedor.html'
    context_object_name = 'object_list'
    paginate_by = int(Configuracion.objects.get(llave = "paginado").valor) 
    
    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")
        if query:
            queryset = queryset.filter(identificador__icontains=query) | queryset.filter(factura__icontains=query) 
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get("q", "")
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Contenedor', 'url': '/contenedor/'},
        ]
        context['breadcrumb'] = breadcrumb
        page_range = range(1, min(5, int(Configuracion.objects.get(llave = "num_pages").valor)) + 1)
        context['page_range'] = page_range
        return context
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} actualizado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido actualizar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
        
class ContenedorCreateView(LoginRequiredMixin, CreateView):
    model = Contenedor
    form_class = ContenedorForm
    template_name = 'contenedor/crear_contenedor.html'
    success_url = reverse_lazy('contenedor_listado')

    def post(self, request, *args, **kwargs):
        form = self.form_class(data = request.POST,files = request.FILES)
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('contenedor_listado')
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
               
class ContenedorUpdateView(LoginRequiredMixin,UpdateView):
    model = Contenedor
    template_name = 'contenedor/actualizar_contenedor.html'
    form_class = ContenedorForm
    success_url = reverse_lazy('contenedor_listado')    
       
class ContenedorDeleteView(LoginRequiredMixin,DeleteView):
    model = Contenedor
    template_name = 'contenedor/eliminar_contenedor.html'
    success_url = reverse_lazy('contenedor_listado')

    def delete(self, request, pk, *args, **kwargs):
        
        contendor = self.get_object()
        contendor.estado = False
        contendor.save()
        mensaje = f'{self.model.__name__} eliminado correctamente!'
        error = 'No hay error!'
        response = JsonResponse({'mensaje': mensaje, 'error': error})
        response.status_code = 201
        return redirect('contenedor_listado')

class LLenarContenedorView(LoginRequiredMixin, DetailView):
    model = Contenedor
    template_name = 'contenedor/llenar_contenedor.html'

    def get_context_data(self, **kwargs):
        contenedor = self.object
        context = super().get_context_data(**kwargs)
        context['contenedor'] = self.object
        context['productos'] = Producto.objects.all()

        total_bultos = 0
        total_bruto = 0
        total_neto = 0
        total_importe = 0
        total_general = 0
        for item in contenedor.items.all():
            importe = item.cantidad * item.precio
            total_bultos += item.bultos
            total_bruto += item.bruto_kg
            total_neto += item.neto_kg
            total_importe += importe    
        total_general = total_importe + contenedor.flete + contenedor.seguro 
        total_bruto = round(total_bruto, 2)
        total_neto = round(total_neto, 2)
        total_importe = round(total_importe, 2)
        
        context['total_bultos'] = total_bultos
        context['total_bruto'] = total_bruto
        context['total_neto'] = total_neto
        context['total_general'] = total_general
        context['total_importe'] = total_importe
        
        breadcrumb = [
            {'text': 'Inicio', 'url': '/home/'},
            {'text': 'Contenedores', 'url': '/contenedor/'},
            {'text': f'{contenedor.id}', 'url': f'/contenedor/llenar/{contenedor.id}/'},
            {'text': 'Detalle', 'url': '/detalle'},
        ]
        context['breadcrumb'] = breadcrumb
        return context

    def post(self, request, *args, **kwargs):
        form = LLenarContenedorForm(data=request.POST)  # Crear una instancia del formulario con los datos POST
        form.instance.contenedor_id = self.kwargs['pk']
        if form.is_valid():
            form.save()
            mensaje = f'{LLenarContenedor.__name__} registrado correctamente!'
            error = 'No hay error!'
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 201
            return redirect('contenedor_llenar', pk=self.kwargs['pk'])
        else:
            mensaje = f'{LLenarContenedor.__name__} no se ha podido registrar!'
            error = form.errors
            response = JsonResponse({'mensaje': mensaje, 'error': error})
            response.status_code = 400
            return response
    
class AgregarProductoCreateView(LoginRequiredMixin, CreateView):
    model = LLenarContenedor
    form_class = LLenarContenedorForm
    context_object_name = 'contenedor'
    template_name = 'contenedor/agregar_producto.html'
    success_url = reverse_lazy('contenedor_listado')
    
    def form_valid(self, form):
        # Guardar los datos del formulario en la tabla de productos
        producto = form.cleaned_data['productos']
        cantidad = form.cleaned_data['cantidad']
        precio = form.cleaned_data['precio']
        bruto_kg = form.cleaned_data['bruto_kg']
        neto_kg = form.cleaned_data['neto_kg']
        bultos = form.cleaned_data['bultos']
        
        # Crea una instancia de LLenarContenedor y guarda los datos en la base de datos
        contenedor = form.save(commit=False)
        contenedor.producto = producto
        contenedor.cantidad = cantidad
        contenedor.precio = precio
        contenedor.bruto_kg = bruto_kg
        contenedor.neto_kg = neto_kg
        contenedor.bultos = bultos
        contenedor.save()
        
        # Si necesitas realizar alguna otra acción adicional, puedes hacerlo aquí
        
        # Llama al método form_valid() de la clase padre para finalizar el proceso de validación del formulario
        return super().form_valid(form)
 
class ActualizarProductoDeleteView(LoginRequiredMixin,UpdateView):
    model = LLenarContenedor
    form_class = LLenarContenedorForm
    template_name = 'contenedor/act_prod_contenedor.html'
    success_url = reverse_lazy('contenedor_llenar')  # URL de éxito

    # Sobrescribir el método get_object para obtener el objeto correcto basado en el ID pasado en la URL
    def get_object(self, queryset=None):
        pk = self.kwargs.get('pk')  # Obtener el ID del contenedor
        llenar_contenedor = LLenarContenedor.objects.get(pk=pk)  # Obtener el objeto LLenarContenedor
        return llenar_contenedor
    
    def get_success_url(self):
        # Obtener el pk del contenedor del producto eliminado
        contenedor_pk = self.object.contenedor.pk
        # Generar la URL inversa para redirigir a contenedor_llenar con el pk del contenedor
        return reverse_lazy('contenedor_llenar', kwargs={'pk': contenedor_pk}) 
    
class EliminarProductoDeleteView(LoginRequiredMixin,DeleteView):
    model = LLenarContenedor
    template_name = 'contenedor/eliminar_prod_contenedor.html'  # Template para confirmar la eliminación
    success_url = reverse_lazy('contenedor_llenar')  # URL de éxito

    # Sobrescribir el método get_object para obtener el objeto correcto basado en el ID pasado en la URL
    def get_object(self, queryset=None):
        pk = self.kwargs.get('pk')  # Obtener el ID del contenedor
        llenar_contenedor = LLenarContenedor.objects.get(pk=pk)  # Obtener el objeto LLenarContenedor
        return llenar_contenedor
    def get_success_url(self):
        # Obtener el pk del contenedor del producto eliminado
        contenedor_pk = self.object.contenedor.pk
        # Generar la URL inversa para redirigir a contenedor_llenar con el pk del contenedor
        return reverse_lazy('contenedor_llenar', kwargs={'pk': contenedor_pk})

""" Generar PDF y EXCEL """
class FacturaView(TemplateView):
    template_name = 'factura.html'

class Factura(LoginRequiredMixin,View):
    model = Contenedor
    success_url = reverse_lazy('contenedor_listado')
    
    def get(self, request, *args, **kwargs):
        productos = Producto.objects.all()
        contenedor_id = kwargs.get('pk')
        contenedor = Contenedor.objects.get(pk=contenedor_id)
        nombre_factura = f"Factura_{contenedor.factura}_{contenedor.cliente}.pdf"
        
        domain = settings.SITE_DOMAIN
        http = settings.HTTPS
        logo_url = f"{http}://{domain}{contenedor.empresa.logo_empresa.url}"
        firma_url = f"{http}://{domain}{contenedor.empresa.firma_empresa.url}"
        
        productos_contenedor = []
        total_bultos = 0
        total_bruto = 0
        total_neto = 0
        total_importe = 0
        total_general = 0
        for item in contenedor.items.all():
            importe = item.cantidad * item.precio  # Calcula el importe de cada producto
            productos_contenedor.append({
                'producto': item.productos,
                'cantidad': item.cantidad,
                'precio': item.precio,
                'importe': importe,
                'bruto_kg': item.bruto_kg,
                'neto_kg': item.neto_kg,
                'bultos': item.bultos,
                
            })
            total_bultos += item.bultos
            total_bruto += item.bruto_kg
            total_neto += item.neto_kg
            total_importe += importe    
        total_general = total_importe + contenedor.flete + contenedor.seguro 
        total_bruto = round(total_bruto,2)
        total_neto = round(total_neto,2)
        total_importe = round(total_importe,2)

        template = get_template('imprimir/factura.html')
        context = {'contenedor': contenedor,
                    'productos': productos,
                    'productos_contenedor': productos_contenedor, 'total_bultos': total_bultos,
                    'total_bruto': total_bruto,
                    'total_neto': total_neto,
                    'total_general': total_general,
                    'total_importe': total_importe,
                    'logo_url': logo_url,
                    'firma_url': firma_url,
                    
                }
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition']= f'attachment; filename="{nombre_factura}"'
        pisa_status = pisa.CreatePDF(
        html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response
    
class Declaracion(LoginRequiredMixin,View):
    model = Contenedor
    success_url = reverse_lazy('contenedor_listado')
    

    def get(self, request, *args, **kwargs):
        
        contenedor_id = kwargs.get('pk')
        contenedor = Contenedor.objects.get(pk=contenedor_id)
        nombre_declaracion = f"Declaración jurada_{contenedor.factura}_{contenedor.cliente}.pdf"
        #logo_url = request.build_absolute_uri(contenedor.empresa.logo_empresa.url)
        #firma_url = request.build_absolute_uri(contenedor.empresa.firma_empresa.url)
        
        domain = settings.SITE_DOMAIN
        http = settings.HTTPS
        logo_url = f"{http}://{domain}{contenedor.empresa.logo_empresa.url}"
        firma_url = f"{http}://{domain}{contenedor.empresa.firma_empresa.url}"
        
        total_importe = sum(item.cantidad * item.precio for item in contenedor.items.all())
        total_general = total_importe + contenedor.flete + contenedor.seguro
        
        template = get_template('imprimir/declaracion_jurada.html')
        context = {'contenedor': contenedor,
                   'logo_url': logo_url,
                   'firma_url': firma_url,
                   'total_general': total_general, 
            }
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition']= f'attachment; filename="{nombre_declaracion}.pdf"'
        pisa_status = pisa.CreatePDF(
        html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

class ExportarContenedorExcelView(LoginRequiredMixin,View):
    def get(self, request, pk):
        # Obtener el contenedor y los productos asociados
        contenedor = Contenedor.objects.get(pk=pk)
        items = LLenarContenedor.objects.filter(contenedor=contenedor)
        nombre_excel = f"Factura_{contenedor.factura}_{contenedor.cliente}.xlsx"
        # Crear un nuevo libro de Excel y seleccionar la hoja activa
        wb = Workbook()
        ws = wb.active

        
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=13)
        ws['B1'] = "Datos Generales del Contenedor"
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].font = Font(size=20, bold=True) 
        ws['B3'] = "No. de Contenedor"
        ws['B4'] = "Proveedor"
        ws['B5'] = "Cadena"
        ws['B6'] = "Cliente"
        ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=11)
        ws['J3'] = "Fecha:"
        ws.merge_cells(start_row=4, start_column=10, end_row=5, end_column=11)
        ws['J4'] = "No. de Factura:"
        ws.merge_cells(start_row=6, start_column=10, end_row=6, end_column=11)
        ws['J6'] = "No. de Contrato:"
        ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=13)
        ws['B7'] = ""

        ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=4)
        ws['B8'] = "Condición de entrega" 
        ws.merge_cells(start_row=8, start_column=5, end_row=8, end_column=7)
        ws['E8'] = "No. BL"
        ws.merge_cells(start_row=8, start_column=8, end_row=8, end_column=9)
        ws['H8'] = "Moneda"
        ws.merge_cells(start_row=8, start_column=10, end_row=8, end_column=11)
        ws['J8'] = "Flete"
        ws.merge_cells(start_row=8, start_column=12, end_row=8, end_column=13)
        ws['L8'] = "Seguro"
        ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=13)
        ws['B10'] = ""

        ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=3)
        ws['B11'] = "Peso Neto Total"
        ws.merge_cells(start_row=11, start_column=4, end_row=11, end_column=5)
        ws['D11'] = "Peso Bruto Total"
        ws.merge_cells(start_row=11, start_column=6, end_row=11, end_column=7)
        ws['F11'] = "Total de Bultos"
        ws.merge_cells(start_row=11, start_column=8, end_row=11, end_column=10)
        ws['H11'] = "Importe FOB"
        ws.merge_cells(start_row=11, start_column=11, end_row=11, end_column=13)
        ws['K11'] = f"Importe {contenedor.transporte}"
        ws.merge_cells(start_row=13, start_column=2, end_row=14, end_column=13)
        ws['B13'] = "Productos del Contenedor"
        ws['B13'].alignment = Alignment(horizontal='center', vertical='center') 
        ws['B13'].font = Font(size=20, bold=True)
        

        

        # Datos del contenedor
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=9)
        ws['C3'] = contenedor.identificador
        ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=9)
        ws['C4'] = contenedor.proveedor.nombre_proveedor
        ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=9)
        ws['C5'] = contenedor.importadora.nombre_importadora
        ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=9)
        ws['C6'] = contenedor.cliente.nombre_cliente
        ws.merge_cells(start_row=3, start_column=12, end_row=3, end_column=13)
        ws['L3'] = contenedor.fecha_creacion.strftime('%Y-%m-%d')
        ws.merge_cells(start_row=4, start_column=12, end_row=5, end_column=13)
        ws['L4'] = contenedor.factura  
        ws.merge_cells(start_row=6, start_column=12, end_row=6, end_column=13)
        ws['L6'] = contenedor.importadora.contrato 

        ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=4)
        ws['B9'] = f"{contenedor.transporte.nombre_transporte}-Incoterm 2010"
        ws.merge_cells(start_row=9, start_column=5, end_row=9, end_column=7)
        ws['E9'] = contenedor.hbl  
        ws.merge_cells(start_row=9, start_column=8, end_row=9, end_column=9)
        ws['H9'] = "USD"
        ws.merge_cells(start_row=9, start_column=10, end_row=9, end_column=11)
        ws['J9'] = contenedor.flete
        ws.merge_cells(start_row=9, start_column=12, end_row=9, end_column=13)
        ws['L9'] = contenedor.seguro

        

        # Encabezados para los productos
        ws['B15'] = "Partida Arancelaria"
        ws.merge_cells(start_row=15, start_column=3, end_row=15, end_column=6)
        ws['C15'] = "Descripción del producto"
        ws['G15'] = "Origen"
        ws['H15'] = "Cantidad"
        ws['I15'] = "Precio"
        ws['J15'] = "Importe"
        ws['K15'] = "Peso Neto"
        ws['L15'] = "Peso Bruto"
        ws['M15'] = "Bultos"

        # Datos de los productos asociados al contenedor
        total_bultos = 0
        neto_total = 0
        bruto_total = 0
        importe_total = 0

        row_num = 16
        for item in items:
            ws.cell(row=row_num, column=2).value = item.productos.partida_arancelaria.partida
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=6)
            ws.cell(row=row_num, column=3).value = item.productos.descripcion
            ws.cell(row=row_num, column=7).value = item.contenedor.puerto_origen.pais_puerto.sigla
            ws.cell(row=row_num, column=8).value = item.cantidad
            ws.cell(row=row_num, column=9).value = item.precio
            ws.cell(row=row_num, column=10).value = item.cantidad * item.precio  
            ws.cell(row=row_num, column=11).value = item.neto_kg
            ws.cell(row=row_num, column=12).value = item.bruto_kg
            ws.cell(row=row_num, column=13).value = item.bultos

            total_bultos += item.bultos
            neto_total += item.neto_kg
            bruto_total += item.bruto_kg
            importe_total += item.cantidad * item.precio
            
            row_num += 1
        
        row_num_1 = row_num-1
        ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=3)
        ws['B12'] = neto_total
        ws.merge_cells(start_row=12, start_column=4, end_row=12, end_column=5)
        ws['D12'] = bruto_total
        ws.merge_cells(start_row=12, start_column=6, end_row=12, end_column=7)
        ws['F12'] = total_bultos
        ws.merge_cells(start_row=12, start_column=8, end_row=12, end_column=10)
        ws['H12'] = importe_total
        ws.merge_cells(start_row=12, start_column=11, end_row=12, end_column=13)
        ws['K12'] = contenedor.flete + contenedor.seguro + importe_total
        

        row_num_2 = row_num+1
        
        ws.merge_cells(start_row=row_num_2, start_column=11, end_row=row_num_2, end_column=11)
        ws.cell(row=row_num_2, column=11).value = "Flete"
        ws.cell(row=row_num_2, column=11).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row_num_2, column=11).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws.merge_cells(start_row=row_num_2, start_column=12, end_row=row_num_2, end_column=13)
        ws.cell(row=row_num_2, column=12).value = contenedor.flete
        ws.cell(row=row_num_2, column=12).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num_2, column=12).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(row=row_num_2, column=13).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        # Seguro
        row_num_3 = row_num_2 + 1
        ws.merge_cells(start_row=row_num_3, start_column=11, end_row=row_num_3, end_column=11)
        ws.cell(row=row_num_3, column=11).value = "Seguro"
        ws.cell(row=row_num_3, column=11).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row_num_3, column=11).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws.merge_cells(start_row=row_num_3, start_column=12, end_row=row_num_3, end_column=13)
        ws.cell(row=row_num_3, column=12).value = contenedor.seguro
        ws.cell(row=row_num_3, column=12).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num_3, column=12).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(row=row_num_3, column=13).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        # Importe Total
        row_num_4 = row_num_3 + 1
        ws.merge_cells(start_row=row_num_4, start_column=11, end_row=row_num_4, end_column=11)
        ws.cell(row=row_num_4, column=11).value = f"Importe Total {contenedor.transporte}"
        ws.cell(row=row_num_4, column=11).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row_num_4, column=11).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        ws.merge_cells(start_row=row_num_4, start_column=12, end_row=row_num_4, end_column=13)
        ws.cell(row=row_num_4, column=12).value = contenedor.flete + contenedor.seguro + importe_total
        ws.cell(row=row_num_4, column=12).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num_4, column=12).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        ws.cell(row=row_num_4, column=13).border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        
        for col in ws.iter_cols(min_row=3, max_row=row_num_4, min_col=2, max_col=13):
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[column].width = adjusted_width

        

        
        for row in ws.iter_rows(min_row=3, max_row=12, min_col=2, max_col=13):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                if isinstance(cell.value, str):
                    cell.font = Font(size=12)

        for row in ws.iter_rows(min_row=15, max_row=row_num_1, min_col=2, max_col=13):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                if isinstance(cell.value, str):
                    cell.font = Font(size=12)

        wb.security = WorkbookProtection(workbookPassword='yu971120*+', lockStructure=True)
        
        # Configurar la respuesta HTTP para devolver el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{nombre_excel}"'
        wb.save(response)
        return response
    
class SolicitarPinView(View):
    template_name = 'proveedor/solicitar_pin.html'

    def get(self, request, nombre_proveedor):
        return render(request, self.template_name, {'nombre_proveedor': nombre_proveedor})

    def post(self, request, nombre_proveedor):
        pin = request.POST.get('pin')
        proveedor = get_object_or_404(Proveedor, nombre_proveedor=nombre_proveedor)

        if proveedor.pin == pin:
            request.session['proveedor_id'] = proveedor.id
            request.session['pin_verified'] = True
            return redirect('contenedores_por_proveedor', nombre_proveedor=nombre_proveedor)
        else:
            return render(request, self.template_name, {'nombre_proveedor': nombre_proveedor, 'error': 'Pin incorrecto'})

class ContenedoresPorProveedorView(View):
    template_name = 'proveedor/contenedor_por_proveedor.html'
    
    def get(self, request, nombre_proveedor):
        proveedor_id = request.session.get('proveedor_id')
        pin_verified = request.session.get('pin_verified', False)
        
        if not proveedor_id or not pin_verified:
            return redirect('solicitar_pin', nombre_proveedor=nombre_proveedor)
        
        
        

        query = request.GET.get('q')
        if query:
            contenedores = Contenedor.objects.filter(proveedor=proveedor_id).filter(
                identificador__icontains=query) | Contenedor.objects.filter(
                factura__icontains=query)
        else:
            contenedores = Contenedor.objects.filter(proveedor=proveedor_id)
            
        proveedor = get_object_or_404(Proveedor, id=proveedor_id, nombre_proveedor=nombre_proveedor)

        paginator = Paginator(contenedores, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'proveedor': proveedor,
            'contenedores': page_obj,
            'q': query,
            'is_paginated': paginator.num_pages > 1,
            'page_obj': page_obj,
            'nombre_proveedor': nombre_proveedor,  # Añadir nombre_proveedor al contexto
        }

        
        
        return render(request, self.template_name, context)
    
class ContenedoresPorProveedorDetalleView(DetailView):
    model = Contenedor
    template_name = 'proveedor/contenedor_por_proveedor_detalle.html'
    context_object_name = 'contenedor'

    def get(self, request, *args, **kwargs):
        nombre_proveedor = kwargs.get('nombre_proveedor')
        proveedor_id = request.session.get('proveedor_id')
        pin_verified = request.session.get('pin_verified', False)

        if not proveedor_id or not pin_verified:
            return redirect('solicitar_pin', nombre_proveedor=nombre_proveedor)
        
        request.session['pin_verified'] = False

        return super().get(request, *args, **kwargs)
    
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contenedor = self.object
        
        # Obtener los productos del contenedor
        context['productos'] = LLenarContenedor.objects.filter(contenedor=contenedor)

        # Calcular los totales
        total_bultos = sum(item.bultos for item in context['productos'])
        total_bruto = sum(item.bruto_kg for item in context['productos'])
        total_neto = sum(item.neto_kg for item in context['productos'])
        total_importe = sum(item.cantidad * item.precio for item in context['productos'])
        total_general = total_importe + contenedor.flete + contenedor.seguro

        context['total_bultos'] = total_bultos
        context['total_bruto'] = total_bruto
        context['total_neto'] = total_neto
        context['total_general'] = total_general
        context['total_importe'] = total_importe

        breadcrumb = [
            {'text': 'Contenedores', 'url': f'/proveedor/{contenedor.proveedor}/contenedores/'},
            {'text': f'{contenedor.id}'},
            {'text': 'Detalle', 'url': '/detalle'},
        ]
        context['breadcrumb'] = breadcrumb
        return context
    
    